import openpyxl

file_path = 'Database.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['職員情報']

# Clear existing data (keep header)
# Header is row 1. Delete all rows from 2 to max_row
if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row - 1)

# New Staff List
staff_data = [
    (1, "船津　勇一"), (2, "長石　庄一郎"), (3, "佐藤　博義"), (4, "裏　久代"), (5, "阿部　映真"),
    (6, "柚木　達也"), (7, "髙倉　圭一"), (8, "安元　正彦"), (9, "宮原　徹"), (10, "堀田　秀俊"),
    (11, "江藤　　賢"), (12, "金谷　昭二"), (13, "橋内　和彦"), (14, "堀谷　桂"), (15, "森　佐和美"),
    (16, "工藤　圭介"), (17, "永松　寛明"), (18, "小川　尚志"), (19, "田北　俊郎"), (20, "堀　竜大"),
    (21, "高田　裕介"), (22, "水江　友和"), (23, "衛藤　加奈"), (24, "西山　幹子"), (25, "藤野　真也"),
    (26, "久保　修平"), (27, "小笠原　陽華"), (28, "工藤　督右"), (29, "藤丸　拓也"), (30, "秦　ひとみ"),
    (31, "三浦　裕希"), (32, "大友　宗一郎"), (33, "秋吉　豊"), (34, "安倍　あいみ"), (35, "伊藤　大貴"),
    (36, "田島　幸太郎"), (37, "古田　義貴"), (38, "佐藤　陽大"), (39, "小池　佑輔"), (40, "樽本　有貴"),
    (41, "中野　豊"), (42, "長野　真結"), (43, "一井　ひろえ"), (44, "玉ノ井　遥廉"), (45, "芳友　麻理子"),
    (46, "佐藤　利佳"), (47, "信岡　大喜"), (48, "川﨑　みゆき"), (49, "中島　貴之"), (50, "奥野　沙耶"),
    (51, "千葉　優希"), (52, "岩永　明日翔"), (53, "梅木　美嶺"), (54, "志賀　真輝"), (55, "首藤　鼓太郎"),
    (56, "野口　大"), (57, "宮﨑　恵子"), (58, "長木　佳子"), (59, "馬場　舜夏"), (60, "轟　睦子"),
    (61, "梶原　希央"), (62, "大曲　直子"), (63, "レミー・ファーニス"), (64, "小溝　晴美"), (65, "宮﨑　晶子"),
    (66, "羽野　紀美子")
]

# Write to sheet
# Cols: 0:id, 1:name, 2:role, 3:order, 4:email, 5:grade, 6:dept1, 7:dept2, 8:dept3, 9:subject, 10:role_type, 11:chief_type
# Excel is 1-indexed. 
# A: ID (auto "01", etc)
# B: Name
# C: Role (Empty)
# D: Order (Int)
# E: Email (Empty)
# ...

for num, name in staff_data:
    # 2-digit ID
    staff_id = f"{num:02d}"
    
    # Row to append
    row_val = [
        staff_id,   # A
        name,       # B
        "",         # C (Role)
        num,        # D (Order)
        "",         # E (Email)
        "",         # F (Grade)
        "",         # G (Dept1)
        "",         # H (Dept2)
        "",         # I (Dept3)
        "",         # J (Subject)
        "",         # K (RoleType)
        ""          # L (ChiefType)
    ]
    ws.append(row_val)

wb.save(file_path)
print(f"Updated staff list with {len(staff_data)} members.")
