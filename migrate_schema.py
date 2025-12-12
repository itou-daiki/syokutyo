import openpyxl
import uuid

file_path = 'Database.xlsx'
wb = openpyxl.load_workbook(file_path)

ws_daily = wb['行事']
# DAILY Cols: A:ID, B:Date, C:Time, D:Content, E:Note, F:EndDate, G:Order, H:ScheduleType
# Index (0-based): 7 is ScheduleType

# 1. Collect Schedule Types from DAILY
date_schedule_map = {}
for row in ws_daily.iter_rows(min_row=2, values_only=False):
    date_val = row[1].value # B
    sch_type = row[7].value # H
    
    if date_val and sch_type:
        date_schedule_map[date_val] = sch_type
    
    # Clear ScheduleType from DAILY
    row[7].value = ""

# 2. Update EVENT Sheet
if 'イベント' not in wb.sheetnames:
    ws_event = wb.create_sheet('イベント')
    ws_event.append(['ID', 'Date', 'Content', 'ScheduleType', 'CleaningStatus'])
else:
    ws_event = wb['イベント']
    # Check header
    header = [c.value for c in ws_event[1]]
    if len(header) < 5:
        ws_event.cell(row=1, column=4).value = 'ScheduleType'
        ws_event.cell(row=1, column=5).value = 'CleaningStatus'

# Helper to find row by date
def find_event_row(date_str):
    for row in ws_event.iter_rows(min_row=2, values_only=False):
        if row[1].value == date_str:
            return row
    return None

# Update existing events and insert new ones
all_target_dates = set(date_schedule_map.keys())

# First pass: Update existing
for row in ws_event.iter_rows(min_row=2, values_only=False):
    d_val = row[1].value
    
    # Set defaults if empty
    # ScheduleType (Col 4 / D)
    if not row[3].value:
        if d_val in date_schedule_map:
            row[3].value = date_schedule_map[d_val]
        else:
            row[3].value = "通常校時"
            
    # CleaningStatus (Col 5 / E)
    if len(row) < 5 or not row[4].value:
        # Ensure cell exists involved
        ws_event.cell(row=row[0].row, column=5).value = "通常清掃"

    if d_val in all_target_dates:
        all_target_dates.remove(d_val)

# Second pass: Create missing events for dates with ScheduleType
for d_val in all_target_dates:
    # A:ID, B:Date, C:Content, D:Type, E:Cleaning
    ws_event.append([
        str(uuid.uuid4()),
        d_val,
        "", # Content empty
        date_schedule_map[d_val],
        "通常清掃"
    ])

# Save
wb.save(file_path)
print("Schema migration complete.")
