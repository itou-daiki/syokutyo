import openpyxl
import uuid

file_path = 'Database.xlsx'
wb = openpyxl.load_workbook(file_path)

# --- DAILY SHEET UPDATE ---
ws_daily = wb['行事']

daily_events = [
    ("2026/02/03", "", "推薦入試検査日", "", "2026/02/03", None),
    ("2026/02/04", "4限", "運委（入試選考会議, 4限）、職員会議（入試号否判定会議, 15:45-16:35, 会議室）", "短縮・清掃カット", "2026/02/04", "短縮校時"),
    ("2026/02/07", "", "大学入学共通テスト模試（2年）", "", "2026/02/07", None),
    ("2026/02/10", "", "第5回入試準備委員会、第7回教務委員会", "", "2026/02/10", None),
    ("2026/02/11", "", "（建国記念の日）", "", "2026/02/11", None),
    ("2026/02/12", "", "SSH成果発表会・STEAM講演会", "", "2026/02/12", None),
    ("2026/02/13", "", "HRA（）学校評議員会（13:30-15:00）、高校入試(一次)願書受付（〜2/19）", "", "2026/02/13", None),
    ("2026/02/17", "", "学年末考査①", "", "2026/02/17", None),
    ("2026/02/18", "", "学年末考査②", "", "2026/02/18", None),
    ("2026/02/19", "", "学年末考査③、高校入試(一次)願書受付（締切日）", "", "2026/02/19", None),
    ("2026/02/20", "", "学年末考査④、サイエンスダイアログ（14:00-15:00）、久大地区高人解研研究大会（13:30〜、会議室）", "", "2026/02/20", None),
    ("2026/02/23", "", "天皇誕生日", "", "2026/02/23", None),
    ("2026/02/24", "", "高校入試(一次)志願変更（〜2/27）", "", "2026/02/24", None),
    ("2026/02/25", "4限", "クラスマッチ（2年）、運委（4限）", "", "2026/02/25", None),
    ("2026/02/26", "", "クラスマッチ（1年）", "", "2026/02/26", None),
    ("2026/02/27", "正午", "（卒業式設営）、素点入力完了（正午）、職員会議（15:15-16:35, 大会議室）、高校入試(一次)志願変更（締切日）", "短縮", "2026/02/27", "短縮校時")
]

print(f"Adding {len(daily_events)} daily events...")

for d in daily_events:
    # A:ID, B:Date, C:Time, D:Content, E:Note, F:EndDate, G:Order, H:ScheduleType
    # d = (Date, Time, Content, Note, EndDate, ScheduleType)
    row = [
        str(uuid.uuid4()), # ID
        d[0],              # Date
        d[1],              # Time
        d[2],              # Content
        d[3],              # Note
        d[4],              # EndDate
        999,               # Order
        d[5] if d[5] else "" # ScheduleType
    ]
    ws_daily.append(row)


# --- EVENT SHEET UPDATE ---
if 'イベント' not in wb.sheetnames:
    ws_event = wb.create_sheet('イベント')
    ws_event.append(['ID', 'Date', 'Content'])
else:
    ws_event = wb['イベント']

main_events = [
    ("2026/02/03", "推薦入試"),
    ("2026/02/11", "建国記念の日"),
    ("2026/02/17", "学年末考査"),
    ("2026/02/18", "学年末考査"),
    ("2026/02/19", "学年末考査"),
    ("2026/02/20", "学年末考査"),
    ("2026/02/23", "天皇誕生日"),
    ("2026/02/25", "2年クラスマッチ"),
    ("2026/02/26", "1年クラスマッチ")
]

print(f"Adding {len(main_events)} main events...")

for m in main_events:
    row = [
        str(uuid.uuid4()),
        m[0],
        m[1]
    ]
    ws_event.append(row)

wb.save(file_path)
print("Done.")
